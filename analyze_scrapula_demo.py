#!/usr/bin/env python3
"""Analyze Scrapula demo file to understand data structure."""

import openpyxl
import json

def analyze_demo_file():
    """Analyze the Scrapula demo Excel file."""
    
    demo_file = "/Users/m4owen/01. Apps/07. Windsurf/03. Claude/DealBot/amazon_products_demo.xlsx"
    
    print("=" * 70)
    print("Scrapula Demo File Analysis")
    print("=" * 70)
    
    try:
        workbook = openpyxl.load_workbook(demo_file)
        
        print(f"\nWorkbook sheets: {workbook.sheetnames}")
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            print(f"\n{'=' * 70}")
            print(f"Sheet: {sheet_name}")
            print('=' * 70)
            
            # Get headers (first row)
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(cell.value)
            
            print(f"\nColumns ({len(headers)}):")
            for i, header in enumerate(headers, 1):
                print(f"  {i}. {header}")
            
            # Show first few rows of data
            print(f"\nFirst 3 rows of data:")
            print("-" * 70)
            
            for row_idx in range(2, min(5, sheet.max_row + 1)):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        row_data[header] = cell_value
                
                print(f"\nRow {row_idx - 1}:")
                for key, value in row_data.items():
                    # Truncate long values
                    value_str = str(value)
                    if len(value_str) > 60:
                        value_str = value_str[:60] + "..."
                    print(f"  {key}: {value_str}")
            
            print(f"\nTotal rows: {sheet.max_row - 1} (excluding header)")
        
        print(f"\n{'=' * 70}")
        print("Field Mapping for DealBot")
        print('=' * 70)
        
        # Create mapping guide
        print(f"\nBased on column names, we need to map:")
        print(f"  ASIN → Look for: 'asin', 'ASIN', 'product_id'")
        print(f"  Title → Look for: 'title', 'product_title', 'name'")
        print(f"  Current Price → Look for: 'price', 'current_price', 'sale_price'")
        print(f"  List Price (PVP) → Look for: 'list_price', 'original_price', 'rrp', 'pvp'")
        print(f"  Image → Look for: 'image', 'image_url', 'main_image'")
        print(f"  Currency → Look for: 'currency', 'currency_code'")
        print(f"  Availability → Look for: 'availability', 'in_stock', 'stock'")
        print(f"  Rating → Look for: 'rating', 'star_rating', 'stars'")
        print(f"  Reviews → Look for: 'reviews', 'review_count', 'ratings_count'")
        
        # Show actual mapping
        print(f"\n{'=' * 70}")
        print("Recommended Field Mapping:")
        print('=' * 70)
        
        sheet = workbook[workbook.sheetnames[0]]
        headers = [cell.value for cell in sheet[1] if cell.value]
        
        mapping = {}
        for header in headers:
            header_lower = header.lower()
            
            if 'asin' in header_lower:
                mapping['asin'] = header
            elif 'title' in header_lower or 'name' in header_lower:
                mapping['title'] = header
            elif 'price' in header_lower and 'list' not in header_lower and 'original' not in header_lower:
                mapping['current_price'] = header
            elif 'list' in header_lower or 'original' in header_lower or 'rrp' in header_lower or 'pvp' in header_lower:
                mapping['list_price'] = header
            elif 'image' in header_lower:
                mapping['image'] = header
            elif 'currency' in header_lower:
                mapping['currency'] = header
            elif 'availability' in header_lower or 'stock' in header_lower:
                mapping['availability'] = header
            elif 'rating' in header_lower or 'star' in header_lower:
                mapping['rating'] = header
            elif 'review' in header_lower:
                mapping['reviews'] = header
        
        print(f"\nDealBot field → Excel column:")
        for field, column in mapping.items():
            print(f"  {field:20} → {column}")
        
        if not mapping:
            print("\n⚠️  Could not auto-detect field mapping")
            print("Please manually review the columns above")
        
        return mapping, headers
        
    except FileNotFoundError:
        print(f"\n❌ File not found: {demo_file}")
        print("Please verify the file path")
        return None, None
    except Exception as e:
        print(f"\n❌ Error reading file: {e}")
        return None, None

if __name__ == "__main__":
    mapping, headers = analyze_demo_file()
    
    if mapping:
        print(f"\n{'=' * 70}")
        print("✅ Ready to update Scrapula service!")
        print('=' * 70)
        print(f"\nNext: I'll update scrapula.py with this field mapping")
