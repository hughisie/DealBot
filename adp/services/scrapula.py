"""Scrapula API service for Amazon product scraping."""

import logging
import requests
import time
import openpyxl
import io
from typing import Optional, List, Dict
from dataclasses import dataclass

logger = logging.getLogger(__name__)


@dataclass
class ScrapulaProductInfo:
    """Product information from Scrapula API."""
    
    asin: str
    title: Optional[str] = None
    image_url: Optional[str] = None
    current_price: Optional[float] = None
    list_price: Optional[float] = None  # PVP/original price
    currency: Optional[str] = None
    availability: Optional[str] = None
    rating: Optional[float] = None
    review_count: Optional[int] = None
    success: bool = False
    error: Optional[str] = None


class ScrapulaService:
    """Service for scraping Amazon product data using Scrapula API."""
    
    def __init__(self, api_key: str, service_name: str = "amazon_product_service"):
        """
        Initialize Scrapula service.
        
        Args:
            api_key: Scrapula API key
            service_name: Scrapula service name for Amazon scraping
        """
        self.api_key = api_key
        self.base_url = "https://api.datapipeplatform.cloud"  # Correct working URL
        self.service_name = service_name
        
        logger.info(f"ScrapulaService initialized with service: {service_name}")
    
    def get_batch_product_data(
        self,
        asins: List[str],
        marketplace: str = "es",
        max_wait_seconds: int = 300
    ) -> Dict[str, ScrapulaProductInfo]:
        """
        Get product data for multiple ASINs using Scrapula batch API.
        
        Args:
            asins: List of Amazon ASINs
            marketplace: Amazon marketplace (es, us, uk, etc.)
            max_wait_seconds: Maximum time to wait for results (default: 5 minutes)
            
        Returns:
            Dict mapping ASIN to ScrapulaProductInfo
        """
        logger.info(f"Fetching batch data for {len(asins)} ASINs from marketplace {marketplace}")
        
        try:
            # Step 1: Create Scrapula task
            task_id = self._create_task(asins, marketplace)
            if not task_id:
                return self._empty_results(asins, "Failed to create Scrapula task")
            
            logger.info(f"Scrapula task created: {task_id}")
            
            # Step 2: Poll for completion
            result = self._wait_for_completion(task_id, max_wait_seconds)
            if not result:
                return self._empty_results(asins, "Task timed out or failed")
            
            # Step 3: Download and parse results
            products = self._parse_results(result, asins)
            
            logger.info(f"Successfully retrieved {len(products)} products from Scrapula")
            return products
            
        except Exception as e:
            error_msg = f"Scrapula batch error: {e}"
            logger.error(error_msg)
            return self._empty_results(asins, error_msg)
    
    def _create_task(self, asins: List[str], marketplace: str) -> Optional[str]:
        """Create a Scrapula task for scraping Amazon products."""
        
        # Map marketplace to Amazon domain
        domain_map = {
            "es": "amazon.es",
            "uk": "amazon.co.uk",
            "us": "amazon.com",
            "ca": "amazon.ca",
            "de": "amazon.de",
            "fr": "amazon.fr",
            "it": "amazon.it",
        }
        
        domain = domain_map.get(marketplace, "amazon.com")
        
        # Create Amazon URLs from ASINs
        queries = [f"https://www.{domain}/dp/{asin}" for asin in asins]
        
        payload = {
            "service_name": self.service_name,
            "queries": queries,
            "settings": {
                "output_extension": "xlsx",
                "marketplace": marketplace
            }
        }
        
        headers = {
            "X-API-KEY": self.api_key,
            "Content-Type": "application/json"
        }
        
        try:
            response = requests.post(
                f"{self.base_url}/tasks",
                headers=headers,
                json=payload,
                timeout=30
            )
            
            response.raise_for_status()
            data = response.json()
            
            return data.get("id")
            
        except requests.exceptions.HTTPError as e:
            logger.error(f"Failed to create Scrapula task: {e}")
            logger.error(f"Response: {e.response.text if e.response else 'No response'}")
            return None
        except Exception as e:
            logger.error(f"Error creating Scrapula task: {e}")
            return None
    
    def _wait_for_completion(self, task_id: str, max_wait_seconds: int) -> Optional[dict]:
        """Poll Scrapula task until completion."""
        
        headers = {"X-API-KEY": self.api_key}
        start_time = time.time()
        poll_interval = 5  # Poll every 5 seconds
        
        while time.time() - start_time < max_wait_seconds:
            try:
                # Use /tasks endpoint with query to get all tasks including completed
                response = requests.get(
                    f"{self.base_url}/tasks?limit=50",
                    headers=headers,
                    timeout=30
                )
                
                response.raise_for_status()
                data = response.json()
                
                # Find our task in the list
                tasks = data.get("tasks", [])
                our_task = None
                for task in tasks:
                    # Match by ui_task_id (short ID from creation response)
                    if task_id in task.get("id", ""):
                        our_task = task
                        break
                
                if our_task:
                    status = our_task.get("status")
                    
                    if status == "SUCCESS":
                        logger.info(f"Scrapula task {task_id} completed successfully")
                        return our_task
                    elif status == "FAILURE":
                        logger.error(f"Scrapula task {task_id} failed")
                        return None
                    else:
                        # Still pending
                        elapsed = int(time.time() - start_time)
                        logger.info(f"Scrapula task still processing... ({elapsed}s)")
                        time.sleep(poll_interval)
                else:
                    # Task not found yet, wait
                    elapsed = int(time.time() - start_time)
                    logger.info(f"Waiting for task to appear... ({elapsed}s)")
                    time.sleep(poll_interval)
                    
            except Exception as e:
                logger.error(f"Error polling Scrapula task: {e}")
                time.sleep(poll_interval)
        
        logger.warning(f"Scrapula task timed out after {max_wait_seconds}s")
        return None
    
    def _parse_results(self, task_result: dict, asins: List[str]) -> Dict[str, ScrapulaProductInfo]:
        """Parse Scrapula task results."""
        
        products = {}
        
        try:
            # Get results - can be file URL or direct data
            results = task_result.get("results", [])
            
            if not results:
                return self._empty_results(asins, "No results in task response")
            
            # Check if results contain file URL
            for result in results:
                if isinstance(result, dict) and "file_url" in result:
                    # Download and parse file
                    file_url = result["file_url"]
                    products = self._download_and_parse_file(file_url, asins)
                    break
                elif isinstance(result, dict):
                    # Direct data in response
                    asin = result.get("asin")
                    if asin:
                        products[asin] = self._parse_response(asin, result)
            
            # Fill in missing ASINs with empty results
            for asin in asins:
                if asin not in products:
                    products[asin] = ScrapulaProductInfo(
                        asin=asin,
                        error="No data returned from Scrapula",
                        success=False
                    )
            
            return products
            
        except Exception as e:
            logger.error(f"Error parsing Scrapula results: {e}")
            return self._empty_results(asins, f"Parse error: {e}")
    
    def _download_and_parse_file(self, file_url: str, asins: List[str]) -> Dict[str, ScrapulaProductInfo]:
        """Download and parse Excel file from Scrapula."""
        
        products = {}
        
        try:
            # Download file
            response = requests.get(file_url, timeout=60)
            response.raise_for_status()
            
            # Parse Excel file
            workbook = openpyxl.load_workbook(io.BytesIO(response.content))
            sheet = workbook[workbook.sheetnames[0]]
            
            # Get headers
            headers = [cell.value for cell in sheet[1] if cell.value]
            
            # Parse each row
            for row_idx in range(2, sheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        row_data[header] = cell_value
                
                asin = row_data.get("asin")
                if asin:
                    products[asin] = self._parse_response(asin, row_data)
            
            return products
            
        except Exception as e:
            logger.error(f"Error downloading/parsing Scrapula file: {e}")
            return {}
    
    def _empty_results(self, asins: List[str], error_msg: str) -> Dict[str, ScrapulaProductInfo]:
        """Create empty results for all ASINs with error message."""
        return {
            asin: ScrapulaProductInfo(asin=asin, error=error_msg, success=False)
            for asin in asins
        }
    
    def get_product_data(self, asin: str, marketplace: str = "es") -> ScrapulaProductInfo:
        """Get single product data (wraps batch method)."""
        results = self.get_batch_product_data([asin], marketplace)
        return results.get(asin, ScrapulaProductInfo(asin=asin, error="Not found", success=False))
    
    def _parse_response(self, asin: str, data: dict) -> ScrapulaProductInfo:
        """
        Parse Scrapula API response.
        
        Args:
            asin: Product ASIN
            data: Response data from API
            
        Returns:
            ScrapulaProductInfo
        """
        # Actual Scrapula API response format:
        # - name: product title
        # - asin: ASIN
        # - price: "$175.00" (string with currency symbol)
        # - strike_price: original price (if discounted)
        # - price_saving: discount amount
        # - availability: stock status
        # - rating: star rating (can be null)
        # - reviews: review count (can be null)
        # - image_1, image_2, etc.: product images
        # - currency: usually null, extract from price string
        
        try:
            # Get title from 'name' field
            title = data.get("name") or data.get("title") or data.get("product_title")
            
            # Parse price string (e.g., "$175.00" or "€63.14")
            price_str = data.get("price", "")
            current_price = self._parse_price(price_str)
            
            # Extract currency from price string
            currency = self._extract_currency_from_price(price_str)
            
            # Get list/strike price if available
            list_price = None
            strike_price_str = data.get("strike_price")
            if strike_price_str:
                list_price = self._parse_price(strike_price_str)
            elif data.get("price_saving"):
                # Calculate from current price + savings
                saving = self._parse_price(data.get("price_saving"))
                if current_price and saving:
                    list_price = current_price + saving
            
            # Get first available image
            image_url = (
                data.get("image_1") or 
                data.get("image_2") or 
                data.get("image_3") or
                data.get("image") or 
                data.get("image_url")
            )
            
            # Parse rating and reviews (can be null)
            rating = self._parse_rating(data.get("rating"))
            review_count = self._parse_int(data.get("reviews"))
            
            product_info = ScrapulaProductInfo(
                asin=asin,
                title=title,
                image_url=image_url,
                current_price=current_price,
                list_price=list_price,
                currency=currency,
                availability=data.get("availability"),
                rating=rating,
                review_count=review_count,
                success=True
            )
            
            logger.info(f"Successfully parsed product {asin}: {product_info.title}")
            return product_info
            
        except Exception as e:
            logger.error(f"Error parsing Scrapula response for {asin}: {e}")
            return ScrapulaProductInfo(
                asin=asin,
                error=f"Parse error: {e}",
                success=False
            )
    
    def _parse_price(self, value) -> Optional[float]:
        """Parse price from various formats."""
        if value is None or value == "":
            return None
        
        try:
            if isinstance(value, (int, float)):
                return float(value)
            
            if isinstance(value, str):
                # Remove currency symbols and commas
                # Handles: "$175.00", "€63.14", "1,234.56", "£99.99"
                clean_value = value.replace(',', '').strip()
                
                # Remove common currency symbols
                for symbol in ['$', '€', '£', '¥', 'USD', 'EUR', 'GBP', 'JPY']:
                    clean_value = clean_value.replace(symbol, '')
                
                clean_value = clean_value.strip()
                if clean_value:
                    return float(clean_value)
                return None
                
        except (ValueError, TypeError):
            return None
    
    def _extract_currency_from_price(self, price_str: str) -> str:
        """Extract currency code from price string."""
        if not price_str:
            return "USD"
        
        # Map symbols to codes
        currency_map = {
            "$": "USD",
            "€": "EUR",
            "£": "GBP",
            "¥": "JPY",
        }
        
        # Check for currency symbols
        for symbol, code in currency_map.items():
            if symbol in price_str:
                return code
        
        # Check for currency codes in string
        for code in ["USD", "EUR", "GBP", "JPY"]:
            if code in price_str.upper():
                return code
        
        # Default to USD
        return "USD"
    
    def _parse_int(self, value) -> Optional[int]:
        """Parse integer from various formats."""
        if value is None:
            return None
        
        try:
            if isinstance(value, str):
                # Remove commas from numbers like "1,234"
                value = value.replace(',', '')
            return int(float(value))
        except (ValueError, TypeError):
            return None
    
    def _parse_currency(self, currency_value) -> str:
        """Parse currency symbol to standard code."""
        if not currency_value:
            return "EUR"
        
        # Map common symbols to codes
        currency_map = {
            "$": "USD",
            "€": "EUR",
            "£": "GBP",
            "¥": "JPY",
        }
        
        # If it's a symbol, convert to code
        if currency_value in currency_map:
            return currency_map[currency_value]
        
        # If already a code (USD, EUR, etc.), return as-is
        return str(currency_value).upper()

    def _parse_rating(self, value) -> Optional[float]:
        """Parse rating from various formats."""
        if value is None:
            return None
        
        try:
            rating = float(value)
            # Ratings should be between 0 and 5
            if 0 <= rating <= 5:
                return rating
            return None
        except (ValueError, TypeError):
            return None
